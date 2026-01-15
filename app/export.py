"""
export.py - Esportazione dati in Excel, CSV e PDF
Production-ready con formatting professionale
"""

from io import BytesIO
from datetime import datetime, date
from decimal import Decimal
from typing import List, Tuple
import logging

logger = logging.getLogger(__name__)


class EsportatorePDF:
    """Esportazione dati trasferte in PDF (.pdf) con design verticale professionale"""

    @staticmethod
    def esporta_trasferte(trasferte: List, nome_file: str = None, data_inizio: str = None, data_fine: str = None, dati_aziendali: dict = None, nome_utente: str = None) -> BytesIO:
        """
        Esporta lista trasferte in PDF orizzontale (A4 landscape) con design professionale
        
        Args:
            trasferte: lista oggetti Trasferta ORM
            nome_file: nome file output (opzionale)
            data_inizio: data inizio periodo (YYYY-MM-DD format)
            data_fine: data fine periodo (YYYY-MM-DD format)
            dati_aziendali: dati aziendali da includere nel footer
            nome_utente: nome completo dell'utente per visualizzazione
            
        Returns:
            BytesIO object con file PDF
        """
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, cm
        except ImportError as ie:
            logger.error(f'reportlab import error: {str(ie)}')
            raise ImportError('Installa: pip install reportlab')

        try:
            output = BytesIO()
            
            # Documento PDF orizzontale (landscape)
            doc = SimpleDocTemplate(
                output,
                pagesize=landscape(A4),
                rightMargin=10,
                leftMargin=10,
                topMargin=25,
                bottomMargin=15
            )

            styles = getSampleStyleSheet()
            
            # Colori professionali
            BRAND_DARK = colors.HexColor('#1a3a52')  # Blu scuro professionale
            BRAND_LIGHT = colors.HexColor('#2c5aa0')  # Blu medio
            ACCENT = colors.HexColor('#0071e3')      # Blu accent
            LIGHT_GRAY = colors.HexColor('#f8f9fa')  # Grigio molto chiaro
            BORDER_GRAY = colors.HexColor('#e0e0e0') # Grigio bordi
            TEXT_DARK = colors.HexColor('#2c3e50')   # Testo scuro
            
            # Stili custom semplici
            header_style = ParagraphStyle(
                'CustomHeader',
                parent=styles['Normal'],
                fontSize=20,
                textColor=BRAND_DARK,
                fontName='Helvetica-Bold',
                spaceAfter=8
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#666666'),
                fontName='Helvetica',
                spaceAfter=2
            )
            
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#555555'),
                fontName='Helvetica'
            )
            
            elements = []

            # ===== HEADER =====
            header_text = 'RIMBORSO KM'
            if nome_utente:
                header_text += f" - {nome_utente}"
            elements.append(Paragraph(header_text, header_style))
            
            # Sottotitolo con periodo intelligente
            if data_inizio and data_fine:
                if data_inizio[:7] == data_fine[:7]:  # Stesso mese
                    # Estrai mese e anno
                    dt = datetime.fromisoformat(data_inizio)
                    mese_nome = ['Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
                                'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre'][dt.month - 1]
                    sottotitolo_text = f"Trasferte di {mese_nome} {dt.year}"
                else:
                    dt_inizio = datetime.fromisoformat(data_inizio)
                    dt_fine = datetime.fromisoformat(data_fine)
                    mesi = ['Gen', 'Feb', 'Mar', 'Apr', 'Mag', 'Giu', 'Lug', 'Ago', 'Set', 'Ott', 'Nov', 'Dic']
                    sottotitolo_text = f"Trasferte da {mesi[dt_inizio.month-1]} {dt_inizio.year} a {mesi[dt_fine.month-1]} {dt_fine.year}"
            else:
                sottotitolo_text = "Report Trasferte Chilometriche"
            
            elements.append(Paragraph(sottotitolo_text, subtitle_style))
            elements.append(Spacer(1, 0.15 * inch))
            
            # Info generali
            data_report = datetime.now().strftime('%d/%m/%Y %H:%M')
            num_trasferte = len(trasferte)
            
            if trasferte:
                data_prima = trasferte[0].data.strftime('%d/%m/%Y')
                data_ultima = trasferte[-1].data.strftime('%d/%m/%Y')
                periodo_text = f"Periodo: {data_prima} → {data_ultima}" if data_prima != data_ultima else f"Data: {data_prima}"
            else:
                periodo_text = "Nessuna trasferta"
            
            info_line = f"Generato: {data_report}  |  {periodo_text}  |  {num_trasferte} trasferte"
            elements.append(Paragraph(info_line, info_style))
            elements.append(Spacer(1, 0.25 * inch))

            # ===== TABELLA TRASFERTE - FULL WIDTH =====
            # Stile per testo dentro celle (per word wrapping)
            cell_style = ParagraphStyle(
                'CellText',
                parent=styles['Normal'],
                fontSize=9,
                fontName='Helvetica',
                textColor=TEXT_DARK,
                leftIndent=3,
                rightIndent=3,
                spaceAfter=0,
                leading=11
            )
            
            data = [['Data', 'Partenza', 'Arrivo', 'Veicolo', 'Km', 'Tariffa €/km', 'Motivo', 'Rimborso €']]

            totale_km = 0.0
            totale_rimborso = 0.0

            for trasferta in trasferte:
                try:
                    km = float(trasferta.chilometri)
                    rimborso = float(trasferta.calcola_rimborso())
                    totale_km += km
                    totale_rimborso += rimborso

                    # Costruisci indirizzi come in Excel: "Nome - Via, Città"
                    partenza = f"{trasferta.nome_partenza or ''} - {trasferta.via_partenza}, {trasferta.citta_partenza}" if trasferta.via_partenza else trasferta.nome_partenza or ''
                    arrivo = f"{trasferta.nome_arrivo or ''} - {trasferta.via_arrivo}, {trasferta.citta_arrivo}" if trasferta.via_arrivo else trasferta.nome_arrivo or ''

                    motivo = (trasferta.motivo or '') if trasferta.motivo else ''
                    
                    # Veicolo e tariffa
                    veicolo_text = ''
                    tariffa_text = ''
                    if trasferta.veicolo:
                        veicolo_text = f"{trasferta.veicolo.marca} {trasferta.veicolo.modello}"
                        tariffa_text = f"{float(trasferta.veicolo.tariffa_km):.4f}"

                    data.append([
                        str(trasferta.data.strftime('%d/%m/%y') if trasferta.data else ''),
                        Paragraph(partenza, cell_style),  # Usa Paragraph per word wrapping
                        Paragraph(arrivo, cell_style),    # Usa Paragraph per word wrapping
                        veicolo_text,
                        f"{km:.1f}",
                        tariffa_text,
                        Paragraph(motivo, cell_style),    # Usa Paragraph per word wrapping
                        f"{rimborso:.2f}"
                    ])
                except Exception as e:
                    logger.warning(f'Errore riga trasferta: {str(e)}')
                    continue

            # Riga totale
            data.append(['TOTALE', '', '', '', f"{totale_km:.1f}", '', f"{num_trasferte} trasferte", f"{totale_rimborso:.2f}"])

            # Crea tabella con colonne proporzionali che coprono tutta la larghezza
            # A4 landscape: 297mm - 20mm margini = 277mm disponibili
            # Conversione: 277mm ≈ 10.9 inches
            table_width = 10.9 * inch
            col_widths = [
                0.6 * inch,   # Data
                1.8 * inch,   # Partenza
                1.8 * inch,   # Arrivo
                1.3 * inch,   # Veicolo
                0.6 * inch,   # Km
                0.8 * inch,   # Tariffa €/km
                2.0 * inch,   # Motivo
                1.0 * inch    # Rimborso
            ]
            
            table = Table(data, colWidths=col_widths, rowHeights=None)  # rowHeights=None permette auto-resize
            
            # Stile tabella professionale
            table.setStyle(TableStyle([
                # ===== HEADER - Blu scuro professionale =====
                ('BACKGROUND', (0, 0), (-1, 0), BRAND_DARK),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('LEFTPADDING', (0, 0), (-1, 0), 5),
                ('RIGHTPADDING', (0, 0), (-1, 0), 5),
                
                # ===== DATI - Layout pulito =====
                ('ALIGN', (0, 1), (0, -2), 'CENTER'),      # Data
                ('ALIGN', (1, 1), (2, -2), 'LEFT'),        # Partenza, Arrivo
                ('ALIGN', (3, 1), (3, -2), 'CENTER'),      # Veicolo
                ('ALIGN', (4, 1), (4, -2), 'RIGHT'),       # Km
                ('ALIGN', (5, 1), (5, -2), 'RIGHT'),       # Tariffa
                ('ALIGN', (6, 1), (6, -2), 'LEFT'),        # Motivo
                ('ALIGN', (7, 1), (7, -2), 'RIGHT'),       # Rimborso
                ('VALIGN', (0, 1), (-1, -2), 'MIDDLE'),
                
                ('FONTSIZE', (0, 1), (-1, -2), 9),
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                ('TEXTCOLOR', (0, 1), (-1, -2), TEXT_DARK),
                ('VALIGN', (0, 1), (-1, -2), 'TOP'),  # Allinea testo in alto per wrapping
                
                # Righe alternate: bianco e grigio molto chiaro
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, LIGHT_GRAY]),
                
                ('TOPPADDING', (0, 1), (-1, -2), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -2), 4),
                ('LEFTPADDING', (0, 1), (-1, -2), 5),
                ('RIGHTPADDING', (0, 1), (-1, -2), 5),
                
                # Grid elegante - linee sottili grigie
                ('GRID', (0, 0), (-1, -1), 0.5, BORDER_GRAY),
                ('LINEBELOW', (0, 0), (-1, 0), 2, BRAND_DARK),  # Linea sotto header
                
                # ===== RIGA TOTALE - Blu chiaro =====
                ('BACKGROUND', (0, -1), (-1, -1), BRAND_LIGHT),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 9),
                ('TOPPADDING', (0, -1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
                ('LEFTPADDING', (0, -1), (-1, -1), 5),
                ('RIGHTPADDING', (0, -1), (-1, -1), 5),
                ('ALIGN', (0, -1), (2, -1), 'LEFT'),
                ('ALIGN', (3, -1), (3, -1), 'RIGHT'),
                ('ALIGN', (4, -1), (4, -1), 'CENTER'),
                ('ALIGN', (5, -1), (5, -1), 'RIGHT'),
            ]))

            elements.append(table)
            elements.append(Spacer(1, 0.25 * inch))
            
            # ===== FOOTER CON DATI AZIENDALI =====
            # Se esistono dati aziendali, mostra nel footer
            if dati_aziendali and dati_aziendali.get('nome_azienda'):
                company_info = dati_aziendali.get('nome_azienda', '')
                details = []
                if dati_aziendali.get('indirizzo_principale'):
                    details.append(dati_aziendali['indirizzo_principale'])
                if dati_aziendali.get('telefono'):
                    details.append(f"Tel. {dati_aziendali['telefono']}")
                if dati_aziendali.get('email'):
                    details.append(dati_aziendali['email'])
                if dati_aziendali.get('partita_iva'):
                    details.append(f"P.IVA: {dati_aziendali['partita_iva']}")
                
                footer_style = ParagraphStyle(
                    'FooterCompany',
                    parent=styles['Normal'],
                    fontSize=7,
                    textColor=colors.HexColor('#666666'),
                    fontName='Helvetica'
                )
                
                elements.append(Paragraph(company_info, footer_style))
                if details:
                    company_details = " • ".join(details)
                    elements.append(Paragraph(company_details, footer_style))
                elements.append(Spacer(1, 0.08 * inch))
            
            # ===== FOOTER ELEGANTE =====
            footer_info = f"KM Totali: {totale_km:.2f}  |  Rimborso Totale: EUR {totale_rimborso:.2f}"
            elements.append(Paragraph(footer_info, info_style))
            elements.append(Spacer(1, 0.08 * inch))
            
            footer_note = "Documento generato automaticamente da Rimborso KM"
            footer_note_style = ParagraphStyle(
                'FooterNote',
                parent=styles['Normal'],
                fontSize=7,
                textColor=colors.HexColor('#999999'),
                fontName='Helvetica'
            )
            elements.append(Paragraph(footer_note, footer_note_style))

            # Build PDF
            logger.info(f'Building professional vertical PDF with {len(trasferte)} trasferte')
            doc.build(elements)
            output.seek(0)
            logger.info('PDF created successfully - Portrait A4 Full Width')

            return output
        
        except Exception as e:
            logger.error(f'PDF generation error: {str(e)}')
            import traceback
            logger.error(traceback.format_exc())
            raise


class EsportatoreExcel:
    """Esportazione dati trasferte in Excel (.xlsx)"""

    @staticmethod
    def esporta_trasferte(trasferte: List, nome_file: str = None, dati_aziendali: dict = None, nome_utente: str = None) -> BytesIO:
        """
        Esporta lista trasferte in Excel con formattazione
        
        Args:
            trasferte: lista oggetti Trasferta ORM
            nome_file: nome file output (default: generato con data)
            dati_aziendali: dati aziendali da includere nell'intestazione
            nome_utente: nome completo dell'utente per visualizzazione
            
        Returns:
            BytesIO object con file Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error('openpyxl non installato')
            raise ImportError('Installa: pip install openpyxl')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Trasferte'

        # Aggiungi dati aziendali come intestazione se presenti
        row_offset = 0
        
        # Aggiungi nome utente
        if nome_utente:
            user_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            user_font = Font(bold=True, size=11)
            
            ws[f'A1'] = f"Utente: {nome_utente}"
            ws[f'A1'].font = user_font
            ws[f'A1'].fill = user_fill
            row_offset = 2
        
        if dati_aziendali and dati_aziendali.get('nome_azienda'):
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            header_font = Font(bold=True, size=12)
            
            ws[f'A{row_offset + 1}'] = dati_aziendali.get('nome_azienda', '')
            ws[f'A{row_offset + 1}'].font = header_font
            ws[f'A{row_offset + 1}'].fill = header_fill
            
            detail_start = row_offset + 2
            details = []
            if dati_aziendali.get('indirizzo_principale'):
                details.append(dati_aziendali['indirizzo_principale'])
            if dati_aziendali.get('telefono'):
                details.append(f"Tel: {dati_aziendali['telefono']}")
            if dati_aziendali.get('email'):
                details.append(f"Email: {dati_aziendali['email']}")
            if dati_aziendali.get('partita_iva'):
                details.append(f"P.IVA: {dati_aziendali['partita_iva']}")
            if dati_aziendali.get('codice_fiscale'):
                details.append(f"Cod.Fis: {dati_aziendali['codice_fiscale']}")
            
            for idx, detail in enumerate(details, start=0):
                ws[f'A{detail_start + idx}'] = detail
                ws[f'A{detail_start + idx}'].font = Font(size=9)
            
            row_offset = detail_start + len(details)

        # Headers
        headers = ['Data', 'Partenza', 'Arrivo', 'Veicolo', 'Km', 'Tariffa €/km', 'Motivo', 'Rimborso (€)']
        header_row = row_offset + 1
        for col, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col, value=header)

        # Styling header
        header_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, 9):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Dati
        totale_rimborso = Decimal('0')
        for idx, trasferta in enumerate(trasferte, start=1):
            rimborso = Decimal(str(trasferta.calcola_rimborso()))
            totale_rimborso += rimborso

            # Costruisci indirizzi completi
            partenza = f"{trasferta.nome_partenza or ''} - {trasferta.via_partenza}, {trasferta.citta_partenza}" if trasferta.via_partenza else trasferta.nome_partenza or ''
            arrivo = f"{trasferta.nome_arrivo or ''} - {trasferta.via_arrivo}, {trasferta.citta_arrivo}" if trasferta.via_arrivo else trasferta.nome_arrivo or ''

            data_row = header_row + idx
            ws.cell(row=data_row, column=1, value=trasferta.data.isoformat() if trasferta.data else '')
            ws.cell(row=data_row, column=2, value=partenza)
            ws.cell(row=data_row, column=3, value=arrivo)
            ws.cell(row=data_row, column=4, value=f"{trasferta.veicolo.marca} {trasferta.veicolo.modello}" if trasferta.veicolo else '')
            ws.cell(row=data_row, column=5, value=float(trasferta.chilometri))
            ws.cell(row=data_row, column=6, value=float(trasferta.veicolo.tariffa_km) if trasferta.veicolo else '')
            ws.cell(row=data_row, column=7, value=trasferta.motivo)
            ws.cell(row=data_row, column=8, value=float(rimborso))

        # Styling dati
        for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + len(trasferte)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Numeri a destra (colonne Km, Tariffa, Rimborso)
            row[4].alignment = Alignment(horizontal='right', vertical='center')
            row[5].alignment = Alignment(horizontal='right', vertical='center')
            row[7].alignment = Alignment(horizontal='right', vertical='center')

            # Formattazione numeri
            row[4].number_format = '0.00'
            row[5].number_format = '0.0000'
            row[7].number_format = '€ #,##0.00'

        # Totale
        total_row = header_row + len(trasferte) + 1
        ws[f'A{total_row}'] = 'TOTALE'
        ws[f'A{total_row}'].font = Font(bold=True, size=11)
        ws[f'H{total_row}'] = float(totale_rimborso)
        ws[f'H{total_row}'].font = Font(bold=True, size=11)
        ws[f'H{total_row}'].number_format = '€ #,##0.00'
        ws[f'H{total_row}'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        # Larghezze colonne
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 15

        # Salva in memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output


class EsportatoreCSV:
    """Esportazione dati trasferte in CSV"""

    @staticmethod
    def esporta_trasferte(trasferte: List) -> str:
        """
        Esporta lista trasferte in CSV
        
        Args:
            trasferte: lista oggetti Trasferta ORM
            
        Returns:
            Stringa CSV
        """
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_ALL)

        # Headers
        writer.writerow([
            'Data', 'Partenza', 'Arrivo', 'Km', 'Metodo', 'Motivo', 'Veicolo', 'Rimborso (€)'
        ])

        # Dati
        for trasferta in trasferte:
            # Costruisci indirizzi completi
            partenza = f"{trasferta.nome_partenza or ''} - {trasferta.via_partenza}, {trasferta.citta_partenza}" if trasferta.via_partenza else trasferta.nome_partenza or ''
            arrivo = f"{trasferta.nome_arrivo or ''} - {trasferta.via_arrivo}, {trasferta.citta_arrivo}" if trasferta.via_arrivo else trasferta.nome_arrivo or ''

            writer.writerow([
                trasferta.data.isoformat() if trasferta.data else '',
                partenza,
                arrivo,
                f"{float(trasferta.chilometri):.2f}",
                trasferta.calcolo_km,
                trasferta.motivo,
                f"{trasferta.veicolo.marca} {trasferta.veicolo.modello}" if trasferta.veicolo else '',
                f"{trasferta.calcola_rimborso():.2f}"
            ])

        return output.getvalue()


def esporta_statistiche(trasferte: List) -> dict:
    """
    Genera statistiche da trasferte per dashboard
    
    Args:
        trasferte: lista oggetti Trasferta ORM
        
    Returns:
        Dict con statistiche giornaliere, mensili, annuali
    """
    stats = {
        'totale_km': 0.0,
        'totale_rimborso': 0.0,
        'numero_trasferte': len(trasferte),
        'media_km_trasferta': 0.0,
        'media_rimborso_trasferta': 0.0,
        'per_data': {},
        'per_mese': {},
        'per_anno': {},
        'per_veicolo': {},
        'per_motivo': {}
    }

    if not trasferte:
        return stats

    # Totali
    for trasferta in trasferte:
        km = float(trasferta.chilometri)
        rimborso = trasferta.calcola_rimborso()
        
        stats['totale_km'] += km
        stats['totale_rimborso'] += rimborso

        # Per data
        data_str = trasferta.data.isoformat()
        if data_str not in stats['per_data']:
            stats['per_data'][data_str] = {'km': 0, 'rimborso': 0, 'count': 0}
        stats['per_data'][data_str]['km'] += km
        stats['per_data'][data_str]['rimborso'] += rimborso
        stats['per_data'][data_str]['count'] += 1

        # Per mese
        mese_key = trasferta.data.strftime('%Y-%m')
        if mese_key not in stats['per_mese']:
            stats['per_mese'][mese_key] = {'km': 0, 'rimborso': 0, 'count': 0}
        stats['per_mese'][mese_key]['km'] += km
        stats['per_mese'][mese_key]['rimborso'] += rimborso
        stats['per_mese'][mese_key]['count'] += 1

        # Per anno
        anno_key = str(trasferta.data.year)
        if anno_key not in stats['per_anno']:
            stats['per_anno'][anno_key] = {'km': 0, 'rimborso': 0, 'count': 0}
        stats['per_anno'][anno_key]['km'] += km
        stats['per_anno'][anno_key]['rimborso'] += rimborso
        stats['per_anno'][anno_key]['count'] += 1

        # Per veicolo
        veicolo_key = f"{trasferta.veicolo.marca} {trasferta.veicolo.modello}"
        if veicolo_key not in stats['per_veicolo']:
            stats['per_veicolo'][veicolo_key] = {'km': 0, 'rimborso': 0, 'count': 0}
        stats['per_veicolo'][veicolo_key]['km'] += km
        stats['per_veicolo'][veicolo_key]['rimborso'] += rimborso
        stats['per_veicolo'][veicolo_key]['count'] += 1

        # Per motivo
        motivo_key = trasferta.motivo
        if motivo_key not in stats['per_motivo']:
            stats['per_motivo'][motivo_key] = {'km': 0, 'rimborso': 0, 'count': 0}
        stats['per_motivo'][motivo_key]['km'] += km
        stats['per_motivo'][motivo_key]['rimborso'] += rimborso
        stats['per_motivo'][motivo_key]['count'] += 1

    # Medie
    if stats['numero_trasferte'] > 0:
        stats['media_km_trasferta'] = round(stats['totale_km'] / stats['numero_trasferte'], 2)
        stats['media_rimborso_trasferta'] = round(stats['totale_rimborso'] / stats['numero_trasferte'], 2)

    # Arrotonda totali
    stats['totale_km'] = round(stats['totale_km'], 2)
    stats['totale_rimborso'] = round(stats['totale_rimborso'], 2)

    return stats
